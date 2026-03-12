import sqlite3
import pandas as pd
from openpyxl import load_workbook

def read_db(dbPath: str):
    connection = sqlite3.connect(dbPath)

    dfVehicleCounts = pd.read_sql_query(
        "SELECT * FROM VehicleCounts", connection
    )

    dfVehicleCounts['timestamp'] = pd.to_datetime(dfVehicleCounts['timestamp'])

    dfGlobalData = pd.read_sql_query(
        "SELECT * FROM GlobalData", connection
    )

    connection.close()

    dfProccessData = pd.DataFrame(
        columns=["Turn", "Vehicle_Type", "Travel_Time", "Quarter_Time"]
    )

    # Dataframe of proccessed data
    for index, row in dfVehicleCounts.iterrows(): 
        turn = row["line"]
        timestamp = row["timestamp"]
        quarterTime = (timestamp.hour * 60 + timestamp.minute) // 15 #NOTE: Start in 0
        diffFrames = row["destination_frame"] - row["origin_frame"]

        dfProccessData.loc[index, "Turn"] = turn
        dfProccessData.loc[index, "Vehicle_Type"] = row["vehicle_type"]
        dfProccessData.loc[index, "Travel_Time"] = diffFrames #TODO: Use fps to change to seconds
        dfProccessData.loc[index, "Quarter_Time"] = quarterTime

    # Dataframe of counts
    dfCounts = dfProccessData.groupby(
        ['Quarter_Time', 'Vehicle_Type']
    ).size().reset_index(name='Counts')

    listVehicleTypes = dfCounts['Vehicle_Type'].values.tolist()
    listTurns = dfProccessData['Turn'].values.tolist()

    #***** CUT SECTION *****#

    dateCount = dfGlobalData['fecha_conteo'][0]
    intersectionName = dfGlobalData['interseccion'][0]
    location = dfGlobalData['ciudad'][0]

    wb = load_workbook('templates/template.xlsx')

    #Inicio writing
    ws = wb['Inicio']
    ws['G5'].value = intersectionName
    ws['G6'].value = dateCount
    ws['G9'].value = location
    
    #Vehicle types
    for i, vehicleType in enumerate(listVehicleTypes):
        ws.cell(4+i,22).value = vehicleType[0].upper() + vehicleType[1:]

    #Turns
    for i, turn in enumerate(listTurns):
        ws.cell(12+i,4).value = turn

    #Conteos
    ws = wb['Conteos']


    print(dfCounts)

    wb.save('./testing/test.xlsx')
    wb.close()

    return None

# if __name__ == '__main__':
#     df = read_db('resources\Conteos.db')