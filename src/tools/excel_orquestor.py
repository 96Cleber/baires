import logging
from openpyxl import load_workbook
import os
import sys
import shutil
import cv2
import math
import pandas as pd
import sqlite3
import json
from tools.outputs.write_counts import write_counts_to_excel
from datetime import datetime

def _timestamp_to_quarter_index(timestamp_str: str) -> int:
    """Convert timestamp string to quarter hour index (1-96).
    
    Args:
        timestamp_str (str): Timestamp string in format "YYYY-MM-DD HH:MM:SS"
    
    Returns:
        int: Quarter hour index (1-96, where 1 is 00:00-00:15, 2 is 00:15-00:30, etc.)
    """
    
    dt = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
    # Calculate minutes from midnight
    minutes_from_midnight = dt.hour * 60 + dt.minute
    # Calculate which 15-minute quarter this falls into (0-95), then add 1 to make it 1-indexed
    quarter_index = (minutes_from_midnight // 15) + 1
    # Ensure it wraps around to 1-96 range (max 96 quarters in a day)
    return ((quarter_index - 1) % 96) + 1

# The old functions _hhmmss_to_seconds and _quarters_covered are no longer needed

def generate_excel_report(db_path: str) -> None:
    """Main function to process database and generate Excel reports.
    
    Args:
        db_path (str): Path to the database file containing all vehicle counts
    """
    # Extract intersection name from db_path for directory structure
    db_dir = os.path.dirname(db_path)
    intersection_name = os.path.splitext(os.path.basename(db_path))[0]
    
    # Create logs directory
    logs_dir = os.path.join(db_dir, "logs")
    os.makedirs(logs_dir, exist_ok=True)
    
    # Configure logging to file
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(os.path.join(logs_dir, "excel_orquestor.log")),
            logging.StreamHandler()  # Also log to console
        ]
    )
    log = logging.getLogger(__name__)
    
    log.info(f"Processing database: {db_path}")
    log.info(f"Intersection name: {intersection_name}")
    
    excel_report_path = os.path.join(db_dir, "reports", f"{intersection_name}.xlsx")
    
    # Creation of report excel
    os.makedirs(os.path.join(db_dir, "reports"), exist_ok=True)
    shutil.copy(
        "./templates/template.xlsx",
        excel_report_path
    )
    log.info(f"Created report template at: {excel_report_path}")

    # Reading typologies
    total_typologies = []
    with open("./templates/tipologias.txt", 'r', encoding='utf-8') as file:
        for line in file:
            clean_line = line.strip()
            if not clean_line.startswith('#') and not clean_line == "":
                total_typologies.append(clean_line)

    # Reading json
    list_files = os.listdir(db_dir)
    json_file = [file for file in list_files if file.endswith('.json')][0]
    with open(os.path.join(db_dir, json_file), 'r') as f:
        json_data = json.load(f)

    total_typologies.remove("Persona") # XXX: Los conteos peatonales irán en otro excel aparte
    
    # Excel workbook load
    wb = load_workbook(excel_report_path, read_only=False, data_only=False)

    # Get intersection info from GlobalData table for filling the Inicio sheet
    # cursor.execute("SELECT fecha_conteo, interseccion, ciudad FROM GlobalData LIMIT 1")
    # global_data = cursor.fetchone()

    # date_count, intersection_name_db, location = global_data
    
    # Fill the Inicio sheet with basic information
    ws_inicio = wb["Inicio"]
    # ws_inicio['G5'].value = intersection_name_db or intersection_name
    # ws_inicio['G6'].value = date_count if date_count else ""
    # ws_inicio['G9'].value = location if location else ""
    
    # Fill vehicle types in the Inicio sheet
    # for i, vehicle_type in enumerate(set(record[1] for record in vehicle_counts)):  # unique vehicle types
    #     ws_inicio.cell(4+i, 22).value = vehicle_type[0].upper() + vehicle_type[1:] if vehicle_type else ""
    
    # Reading json for turns
    direction_dict = {"N": [], "S": [], "E": [], "O": []}
    movements_json = json_data["movements"]
    counting_lines_json = json_data["counting_lines"]
    for mov_dict in movements_json:
        origin_mov = mov_dict["o"]
        for counting_line in counting_lines_json:
            access_line = counting_line["access"]
            main_access_line = access_line[0] if len(access_line) > 1 else access_line
            if origin_mov == counting_line["id"]:
                row_excel_mov = {
                    "origin": origin_mov,
                    "destination": mov_dict["d"],
                    "turn": mov_dict["id"],
                    "access": counting_line["name"]
                }
                direction_dict[main_access_line].append(row_excel_mov)
                break

    # Fill movements in the Inicio sheet
    excel_positions = {
        "N": (12,5),
        "S": (12,11),
        "E": (24,5),
        "O": (24,11)
    }
    
    for direction, cell_range in excel_positions.items():
        row = cell_range[0]
        column = cell_range[1]
        for i, row_dict in enumerate(direction_dict[direction]):
            ws_inicio.cell(row=row+i, column=column).value = str(row_dict["origin"])
            ws_inicio.cell(row=row+i, column=column+1).value = str(row_dict["destination"])
            ws_inicio.cell(row=row+i, column=column+2).value = str(row_dict["turn"])
            ws_inicio.cell(row=row+i, column=column+3).value = str(row_dict["access"])

    wb.save(excel_report_path)
    
    # Connect to the database
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Get all vehicle counts from the database
        cursor.execute("SELECT movement, vehicle_type, origin_frame, destination_frame, timestamp FROM VehicleCounts")
        vehicle_counts = cursor.fetchall()
        
        log.info(f"Found {len(vehicle_counts)} vehicle records in database")
        
        # Group vehicle counts by quarter based on destination_frame timestamp
        # Dictionary to store counts by quarter, movement, and vehicle type
        quarter_data = {}
        for movement, vehicle_type, origin_frame, destination_frame, timestamp in vehicle_counts:
            # Use the destination_frame timestamp to determine the quarter
            quarter_index = _timestamp_to_quarter_index(destination_frame)
            
            if quarter_index not in quarter_data:
                quarter_data[quarter_index] = []
            
            quarter_data[quarter_index].append({
                'movement': movement,
                'vehicle_type': vehicle_type,
                'origin_frame': origin_frame,
                'destination_frame': destination_frame,
                'timestamp': timestamp
            })
        
        log.info(f"Grouped vehicle records into {len(quarter_data)} quarters")
        
        # Process each quarter in the data
        for quarter_index, records in quarter_data.items():
            log.info(f"Processing quarter {quarter_index} with {len(records)} records")
            # Create a temporary DataFrame-like structure for this quarter's data
            # We'll pass this to the write_counts function by simulating a video folder structure
            
            # Create a temporary directory structure to simulate video folder
            temp_video_folder = os.path.join(db_dir, f"temp_{quarter_index}")
            os.makedirs(temp_video_folder, exist_ok=True)
            
            # Create a temporary database for this quarter (to be used by write_counts)
            temp_db_path = os.path.join(temp_video_folder, "Conteos.db")
            temp_conn = sqlite3.connect(temp_db_path)
            temp_cursor = temp_conn.cursor()
            
            # Create the table structure
            temp_cursor.execute('''CREATE TABLE IF NOT EXISTS VehicleCounts (
                                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                                    movement TEXT,
                                    vehicle_type TEXT,
                                    origin_frame INTEGER,
                                    destination_frame INTEGER,
                                    timestamp TEXT
                                  )''')
                                  
            # Insert records for this quarter
            for record in records:
                temp_cursor.execute("INSERT INTO VehicleCounts (movement, vehicle_type, origin_frame, destination_frame, timestamp) VALUES (?, ?, ?, ?, ?)",
                                  (record['movement'], record['vehicle_type'], record['origin_frame'], record['destination_frame'], record['timestamp']))
                                  
            temp_conn.commit()
            temp_conn.close()
            
            # Now call write_counts with the quarter index and the simulated video folder
            write_counts_to_excel(
                temp_video_folder,
                quarter_index + 8,  # NOTE: +8 Because of space in Excel
                wb,
                excel_report_path,
                total_typologies
            )
            
            # Clean up temporary folder after processing
            shutil.rmtree(temp_video_folder, ignore_errors=True)

        log.info("Successfully processed all quarters from database")

    except Exception as e:
        log.error(f"Error processing database: {str(e)}", exc_info=True)
    finally:
        conn.close()
        wb.close()
        log.info("Database connection closed and Excel workbook saved successfully")