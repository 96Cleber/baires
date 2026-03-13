import os
import sys
from openpyxl import load_workbook
from read_db import read_db


def get_resource_path(relative_path: str) -> str:
    """Obtener ruta de recurso, compatible con PyInstaller"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), relative_path)


# Reading dbPath
dbPath = "./resources/data.sqlite"
df, type_string = read_db(dbPath)
print(df)
print(type_string)

# Writing to excel
templatePath = get_resource_path("templates/template.xlsx")
wb = load_workbook(templatePath)

# Writing to excel

# Car types
ws = wb['Inicio']
for i, elem in enumerate(type_string):
    ws.cell(row=4+i, column=22).value = elem.upper()

# Counts by quarter of hour
for direction in ["N", "S", "E", "O"]:
    ws = wb[direction]
    

wb.save("./templates/test.xlsx")
wb.close()